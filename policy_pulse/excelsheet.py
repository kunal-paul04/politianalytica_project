import os
from openpyxl import load_workbook

def get_excel_data():
    file_path = os.path.join(os.getcwd(), 'media', 'data.xlsx')
    wb = load_workbook(file_path)
    sheet = wb.active
    
    data = []
    # Assuming the first row contains headers
    headers = [cell.value for cell in sheet[1]]
    
    # Selective columns to display
    columns_to_display = ['Party', 'Ideology', 'Founded', 'Leader', 'Lokh Sabha Seats', 'Rajya Sabha Seats']
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        selected_row = [row[headers.index(col)] if col in headers else '' for col in columns_to_display]
        data.append(selected_row)
    
    return data